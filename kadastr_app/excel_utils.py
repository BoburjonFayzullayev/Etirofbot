import openpyxl
from .models import KadastrMalumat, ObyektMalumat


# ─── To'lov Excel ustun mapping ───────────────────────────────────────────────
USTUN_MAPPING = {
    'viloyat':        ['viloyat', 'viloyat nomi', 'b'],
    'tuman':          ['tuman', 'tuman nomi', 'c'],
    'mfy':            ['mfy', 'd'],
    'kocha':          ["ko'cha nomi", 'kocha', 'e'],
    'kadastr_raqami': ['kadastr raqami', 'kadastr', 'f'],
    'invoys_raqami':  ['invoys raqami', 'invoys', 'g'],
    'summa_miqdori':  ['summa miqdori', "to'lov miqdori", 'miqdori', 'h'],
    'tolovchi_fio':   ["to'lovchi f.i.o", "to'lovchi", 'fio', 'i'],
    'tolov_holati':   ["to'lov holati", 'holati', 'j'],
}

# ─── Obyekt holati Excel ustun mapping ────────────────────────────────────────
OBYEKT_USTUN_MAPPING = {
    'kadastr_raqami': ['kadastr raqami', 'kadastr', 'a'],
    'viloyat':        ['viloyat nomi', 'viloyat', 'b'],
    'tuman':          ['tuman nomi', 'tuman', 'c'],
    'mfy':            ['mfy', 'd'],
    'holati':         ['holati', 'e'],
}


def _sarlavha_toping(rows):
    """Birinchi bo'sh bo'lmagan qatorni sarlavha sifatida qaytaradi"""
    for i, row in enumerate(rows):
        if any(c.value for c in row):
            return row, i
    return None, 0


def _ustun_indekslar(header_row, mapping):
    """Sarlavha qatoridan ustun indekslarini toping"""
    indekslar = {}
    for cell in header_row:
        if cell.value:
            qiymat = str(cell.value).strip().lower()
            for kalit, variantlar in mapping.items():
                if qiymat in variantlar:
                    indekslar[kalit] = cell.column - 1
                    break
    return indekslar


def _qiymat_ol(row_values, indekslar, kalit, standart=''):
    idx = indekslar.get(kalit)
    if idx is not None and idx < len(row_values):
        val = row_values[idx]
        return str(val).strip() if val is not None else standart
    return standart


# ─── To'lov Excel yuklash ─────────────────────────────────────────────────────

def excel_faylni_o_qi(excel_upload_obj):
    """To'lov Excel faylini o'qib, KadastrMalumat bazaga saqlaydi"""
    wb = openpyxl.load_workbook(excel_upload_obj.fayl.path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        raise ValueError("Excel fayl bo'sh!")

    header_row, header_idx = _sarlavha_toping(rows)
    if not header_row:
        raise ValueError("Sarlavha qatori topilmadi!")

    indekslar = _ustun_indekslar(header_row, USTUN_MAPPING)
    if not indekslar:
        indekslar = {
            'viloyat': 1, 'tuman': 2, 'mfy': 3, 'kocha': 4,
            'kadastr_raqami': 5, 'invoys_raqami': 6,
            'summa_miqdori': 7, 'tolovchi_fio': 8, 'tolov_holati': 9,
        }

    malumatlar = []
    for row in rows[header_idx + 1:]:
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        kadastr = _qiymat_ol(row_values, indekslar, 'kadastr_raqami')
        if not kadastr:
            continue
        malumatlar.append(KadastrMalumat(
            excel_fayl=excel_upload_obj,
            viloyat=_qiymat_ol(row_values, indekslar, 'viloyat'),
            tuman=_qiymat_ol(row_values, indekslar, 'tuman'),
            mfy=_qiymat_ol(row_values, indekslar, 'mfy'),
            kocha=_qiymat_ol(row_values, indekslar, 'kocha'),
            kadastr_raqami=kadastr,
            invoys_raqami=_qiymat_ol(row_values, indekslar, 'invoys_raqami'),
            summa_miqdori=_qiymat_ol(row_values, indekslar, 'summa_miqdori'),
            tolovchi_fio=_qiymat_ol(row_values, indekslar, 'tolovchi_fio'),
            tolov_holati=_qiymat_ol(row_values, indekslar, 'tolov_holati'),
        ))

    KadastrMalumat.objects.bulk_create(malumatlar)
    return len(malumatlar)


# ─── Obyekt holati Excel yuklash ──────────────────────────────────────────────

def obyekt_excel_o_qi(obyekt_upload_obj):
    """Obyekt holati Excel faylini o'qib, ObyektMalumat bazaga saqlaydi"""
    wb = openpyxl.load_workbook(obyekt_upload_obj.fayl.path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))
    if not rows:
        raise ValueError("Excel fayl bo'sh!")

    header_row, header_idx = _sarlavha_toping(rows)
    if not header_row:
        raise ValueError("Sarlavha qatori topilmadi!")

    indekslar = _ustun_indekslar(header_row, OBYEKT_USTUN_MAPPING)
    if not indekslar:
        # Standart: A=0, B=1, C=2, D=3, E=4
        indekslar = {
            'kadastr_raqami': 0,
            'viloyat': 1,
            'tuman': 2,
            'mfy': 3,
            'holati': 4,
        }

    malumatlar = []
    for row in rows[header_idx + 1:]:
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue
        kadastr = _qiymat_ol(row_values, indekslar, 'kadastr_raqami')
        if not kadastr:
            continue
        malumatlar.append(ObyektMalumat(
            excel_fayl=obyekt_upload_obj,
            kadastr_raqami=kadastr,
            viloyat=_qiymat_ol(row_values, indekslar, 'viloyat'),
            tuman=_qiymat_ol(row_values, indekslar, 'tuman'),
            mfy=_qiymat_ol(row_values, indekslar, 'mfy'),
            holati=_qiymat_ol(row_values, indekslar, 'holati'),
        ))

    ObyektMalumat.objects.bulk_create(malumatlar)
    return len(malumatlar)